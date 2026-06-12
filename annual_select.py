# ==============================
# 【年1回実行】長期安定配当株 メイン選定スクリプト（Dモデル・本格版）
# 実行タイミング：年1回（毎年4月1日 ※休日の場合は翌営業日）
# 役割：全プライム上場銘柄をスキャンしてFinal7を決定・保存
#
# 【Dモデルの改良点】
# 改良1: 株価を履歴データから自前取得（info.get()の不安定さを排除）
# 改良2: 配当利回りを直近3年平均配当で計算（単年異常値の影響を平滑化）
# 改良3: ROE・配当性向を複数年平均で計算（単年決算異常値の影響を平滑化）
# 改良4: 異常値はcontinueで除外せずフラグを立てて残す（目視確認可能）
# 改良5: 直近3ヶ月以内の株式分割を検出し「要確認」フラグを付与
#         （分割直後はyfinanceの配当・株価調整がズレて利回り異常値が出やすい）
# 改良6: 安定性スコア計算時に分割逆換算を適用
#         （yfinance調整済み配当をそのまま前年比すると増配が減配に見える問題を解消）
# 改良7: 総合点を90点→100点満点に変更【C案】
#         利回り×2.5（+0.5）、配当性向×2.0（+0.5）で利回り・持続性を強化
#         安定性・ROE・成長・財務は据え置き
# 改良8: DOE・理論利回り・実質PBR倍率を算出・出力
#         DOE=ROE×配当性向/100、理論利回り=DOE/PBR
#         実質PBR倍率=DOE/現実利回り（2倍超は成長期待先行銘柄の目安）
#         利回り×2.5（+0.5）、配当性向×2.0（+0.5）で利回り・持続性を強化
#         安定性・ROE・成長・財務は据え置き
# 改良9: Isolation Forest異常検知結果をExcel Final7シートに色付き警告表示
#         anomaly_flag=-1の銘柄セルを黄色背景・太字でハイライト
#         monthly連続異常（2回連続score≤-0.5）でexclusion_candidate=trueをJSON付与
# ==============================
 
import yfinance as yf
from datetime import datetime
import pandas as pd
from tqdm import tqdm
import os
import json
import numpy as np
import jquantsapi
import math
from sklearn.ensemble import IsolationForest, RandomForestClassifier
from sklearn.preprocessing import RobustScaler
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
 
try:
    from email_config import SMTP_SERVER, SMTP_PORT, EMAIL_ADDRESS, EMAIL_PASSWORD, TO_EMAIL
    MAIL_ENABLED = True
except ImportError:
    MAIL_ENABLED = False
    print("⚠️  email_config が見つかりません。メール通知はスキップします。")
 
# ========= 設定 =========
YEARS      = 15
AVG_YEARS  = 3        # 利回り・ROE・配当性向の平均計算に使う年数
JPX_FILE   = "data_j.xlsx"
OUTPUT_DIR = "output"
ANNUAL_RESULT_FILE = os.path.join(OUTPUT_DIR, "annual_result.xlsx")
FINAL7_JSON        = os.path.join(OUTPUT_DIR, "final7.json")
JQUANTS_API_KEY    = "RtZSrXhB8a2ytDc-iT-Q1zFNz0II1rSj8f9wfREEvrc"  # J-Quants APIキー
 
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ========= ML設定 =========
# Isolation Forestで使用する特徴量
ANOMALY_FEATURES = [
    "利回り%(3年平均)",
    "配当性向%(複数年平均)",
    "ROE%(複数年平均)",
    "DOE%",
    "PBR",
    "理論利回り%",
    "実質PBR倍率",
    "売上成長率%",
    "負債比率",
]
ANOMALY_CONTAMINATION = 0.05   # 異常の想定割合（5%）
FEATURE_IMPORTANCE_OUT = os.path.join(OUTPUT_DIR, "feature_importance.png")
EXPORT_FEATURE_IMPORTANCE = True   # Falseにすると重要度PNG出力をスキップ
 
today = datetime.today()
 
# 4月1日起点の会計年度ベース
if today.month >= 4:
    fiscal_year = today.year
else:
    fiscal_year = today.year - 1
 
start_year = fiscal_year - YEARS
 
 
# ========= メール通知 =========
def send_alert(subject, body):
    if not MAIL_ENABLED:
        print(f"[メール省略] {subject}\n{body}")
        return
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    msg = MIMEMultipart()
    msg["From"]    = EMAIL_ADDRESS
    msg["To"]      = TO_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))
    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
    server.send_message(msg)
    server.quit()
 
 
# ============================================================
# 【改良1】株価を履歴データから自前取得
# 【旧コード】info.get("currentPrice") → 取得ミスが多く異常値の原因
# 【新コード】history()の直近終値を使用、失敗時はinfoにフォールバック
# ============================================================
def get_current_price(stock):
    try:
        hist = stock.history(period="1mo")
        if hist.empty:
            return None
        closes = hist["Close"].dropna()
        if closes.empty:
            return None
        return float(closes.iloc[-1])
    except Exception:
        return None
 
 
# ============================================================
# 【改良3】ROE・配当性向を複数年平均で計算
# 【旧コード】info.get("returnOnEquity") / info.get("payoutRatio") の単年値
# 【新コード】financials/balance_sheet/cashflowから複数年平均を計算
#             取得できない場合はNoneを返しinfoの単年値にフォールバック
# ============================================================
def get_multi_year_financials(stock, avg_years):
    try:
        financials = stock.financials
        balance    = stock.balance_sheet
        cashflow   = stock.cashflow
 
        if financials.empty or balance.empty:
            return None, None
 
        roe_list    = []
        payout_list = []
        cols = financials.columns[:avg_years]
 
        for col in cols:
            try:
                # 純利益
                net_income = None
                for label in ["Net Income", "Net Income Common Stockholders"]:
                    if label in financials.index:
                        net_income = financials.loc[label, col]
                        break
 
                # 自己資本
                equity = None
                for label in ["Stockholders Equity", "Total Stockholders Equity",
                              "Common Stock Equity"]:
                    if label in balance.index:
                        equity = balance.loc[label, col]
                        break
 
                # 配当総額
                div_paid = None
                if not cashflow.empty and col in cashflow.columns:
                    for label in ["Cash Dividends Paid", "Common Stock Dividend Paid"]:
                        if label in cashflow.index:
                            div_paid = abs(cashflow.loc[label, col])
                            break
 
                if net_income and equity and equity != 0:
                    roe_list.append(net_income / equity * 100)
 
                if div_paid and net_income and net_income > 0:
                    payout_list.append(div_paid / net_income * 100)
 
            except Exception:
                continue
 
        avg_roe    = float(np.mean(roe_list))    if roe_list    else None
        avg_payout = float(np.mean(payout_list)) if payout_list else None
        return avg_roe, avg_payout
 
    except Exception:
        return None, None
 
 
# ============================================================
# 【改良5】株式分割検出（J-Quants版）
# yfinanceの分割データは異常値（極小値）が混入するケースがあるため
# J-Quants APIの公式データ（AdjFactor）で代替する。
# AdjFactorは権利落ち日に「1/分割比率」が記録される。
# 例）3分割 → 0.333333、5分割 → 0.2
# 上場廃止銘柄はJ-Quantsで取得できないため、その場合はyfinanceにフォールバック。
# ============================================================
SPLIT_WATCH_MONTHS = 3    # 何ヶ月以内の分割を警戒対象にするか
 
# J-Quantsクライアントを初期化（グローバルで1回だけ）
try:
    _jquants_client = jquantsapi.ClientV2(api_key=JQUANTS_API_KEY)
except Exception:
    _jquants_client = None
 
def check_recent_split(stock, months=SPLIT_WATCH_MONTHS):
    """
    直近 months ヶ月以内に株式分割があれば (True, 分割比率の文字列) を返す。
    なければ (False, "") を返す。
    J-QuantsのAdjFactorを優先使用し、取得失敗時はyfinanceにフォールバック。
    """
    # --- J-Quants版（優先） ---
    if _jquants_client is not None:
        try:
            code = stock.ticker.replace(".T", "")
            df = _jquants_client.get_eq_bars_daily(code=code)
            if df is not None and not df.empty:
                df['Date'] = pd.to_datetime(df['Date'])
                cutoff = pd.Timestamp.now() - pd.DateOffset(months=months)
                recent = df[df['Date'] >= cutoff]
                split_rows = recent[recent['AdjFactor'] != 1.0]
                if not split_rows.empty:
                    notes = []
                    for _, row in split_rows.iterrows():
                        ratio = round(1.0 / row['AdjFactor'], 1)
                        notes.append(f"{ratio:.1f}倍({row['Date'].strftime('%Y-%m-%d')})")
                    return True, " / ".join(notes)
                return False, ""
        except Exception:
            pass
 
    # --- yfinanceフォールバック ---
    try:
        splits = stock.splits
        if splits.empty:
            return False, ""
        cutoff = pd.Timestamp.now(tz="UTC") - pd.DateOffset(months=months)
        recent = splits[splits.index >= cutoff]
        if recent.empty:
            return False, ""
        notes = []
        for dt, ratio in recent.items():
            if ratio < 0.01:
                continue
            notes.append(f"{ratio:.1f}倍({dt.strftime('%Y-%m-%d')})")
        if not notes:
            return False, ""
        return True, " / ".join(notes)
    except Exception:
        return False, ""
 
def normalize_div_for_stability(yearly_div: dict, stock) -> dict:
    """
    yearly_div（yfinance調整済み）を分割前名目ベースに逆換算して返す。
    splits データが取得できない場合は元の yearly_div をそのまま返す。
    """
    try:
        splits = stock.splits
        if splits.empty:
            return yearly_div
 
        # 年ごとの累積分割比率を計算
        # splits は「その日以降に適用された分割比率」の時系列
        # yfinance は全過去データを「最新の分割後ベース」に換算済みなので、
        # ある年Yの配当を名目に戻すには「Y年以降に起きた全分割比率の積」を掛ける
        split_years = {}
        for dt, ratio in splits.items():
            y = dt.year
            split_years[y] = split_years.get(y, 1.0) * float(ratio)
 
        all_years = sorted(set(list(yearly_div.keys()) + list(split_years.keys())))
 
        # 各年について「その年以降（その年を含む）の累積分割比率」を計算
        cumulative = {}
        running = 1.0
        for y in reversed(all_years):
            running *= split_years.get(y, 1.0)
            cumulative[y] = running
 
        # 最古年より前の分割は running に既に含まれているので
        # 最小年の cumulative をベースに各年の補正係数を決める
        min_year = min(all_years)
        base = cumulative.get(min_year, 1.0)
 
        normalized = {}
        for y, div in yearly_div.items():
            # その年以降の累積分割比率（その年の分割は「その年の期末配当後」に
            # 実施されると仮定し、当該年には適用しない）
            # → cumulative[y] にはy年の分割が含まれるので y+1 以降の比率を使う
            after_ratio = 1.0
            for sy, sr in split_years.items():
                if sy > y:
                    after_ratio *= sr
            normalized[y] = div * after_ratio
 
        return normalized
 
    except Exception:
        return yearly_div
 
 
# ========= 配当安定性スコア（増配・減配耐性） =========
def calc_dividend_stability(yearly_div, start_year, end_year):
    years = list(range(start_year, end_year))
    divs  = [yearly_div.get(y, 0) for y in years]
 
    increase = decrease = maintain = 0
    for i in range(1, len(divs)):
        prev, curr = divs[i-1], divs[i]
        if prev <= 0:
            continue
        rate = (curr - prev) / prev
        if rate > 0.01:
            increase += 1
        elif rate < -0.01:
            decrease += 1
        else:
            maintain += 1
 
    total = increase + decrease + maintain
    if total == 0:
        return 5.0
    score = (increase * 1.0 + maintain * 0.5 - decrease * 2.0) / total
    return round(max(0.0, min(10.0, score * 10)), 2)
 
 
# ========= 連続値スコア（0〜10点） =========
def calc_score(value, low, high, reverse=False):
    if value is None:
        return 0.0
    if reverse:
        value, low, high = -value, -high, -low
    score = (value - low) / (high - low) * 10
    return round(max(0.0, min(10.0, score)), 2)
 
 
# ============================================================
# 【追加】Isolation Forest による異常検知
# 対象：データ品質「正常」銘柄のみ
# 出力列：anomaly_flag（-1=異常, 1=正常）、anomaly_score（低いほど異常度高）
#         anomaly_reason（スコアが低い主因列を列挙）
# ============================================================
def detect_anomalies(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["anomaly_flag"]   = 0
    df["anomaly_score"]  = 0.0
    df["anomaly_reason"] = ""

    target = df[df["データ品質"] == "正常"].copy()
    if len(target) < 10:
        print("⚠️  異常検知: 正常銘柄が少なすぎるためスキップ")
        return df

    # 使用可能な列のみに絞る（全NaN列を除外）
    use_cols = [c for c in ANOMALY_FEATURES if c in target.columns
                and target[c].notna().sum() > 0]
    X_raw = target[use_cols].fillna(0)

    # RobustScaler（外れ値に強い正規化）
    scaler = RobustScaler()
    X_scaled = scaler.fit_transform(X_raw)

    iso = IsolationForest(
        n_estimators=300,
        contamination=ANOMALY_CONTAMINATION,
        random_state=42,
        n_jobs=-1,
    )
    flags  = iso.fit_predict(X_scaled)   # -1 or 1
    scores = iso.score_samples(X_scaled) # 低いほど異常

    df.loc[target.index, "anomaly_flag"]  = flags
    df.loc[target.index, "anomaly_score"] = np.round(scores, 4)

    # 異常銘柄に対して「どの列が外れているか」を簡易列挙
    X_scaled_df = pd.DataFrame(X_scaled, index=target.index, columns=use_cols)
    anomaly_idx = target.index[flags == -1]
    for idx in anomaly_idx:
        row_abs = X_scaled_df.loc[idx].abs()
        top = row_abs.nlargest(3).index.tolist()
        df.loc[idx, "anomaly_reason"] = " / ".join(top)

    n_anomaly = int((flags == -1).sum())
    print(f"✅  異常検知完了: {n_anomaly} 銘柄を異常判定 (全{len(target)}社中)")
    return df


# ============================================================
# 【追加】RandomForest による特徴量重要度の可視化
# ラベル：総合点上位30%を1、下位30%を0として疑似ラベル生成
# （教師あり学習のラベルがなくても重要度の傾向を把握できる）
# ============================================================
def export_feature_importance(df: pd.DataFrame) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm

        target = df[df["データ品質"] == "正常"].copy()
        use_cols = [c for c in ANOMALY_FEATURES if c in target.columns
                    and target[c].notna().sum() > 0]
        if len(target) < 20 or len(use_cols) < 3:
            print("⚠️  特徴量重要度: データ不足のためスキップ")
            return

        X = target[use_cols].fillna(0)
        threshold_high = target["総合点"].quantile(0.70)
        threshold_low  = target["総合点"].quantile(0.30)
        mask = (target["総合点"] >= threshold_high) | (target["総合点"] <= threshold_low)
        X_sub = X[mask]
        y_sub = (target.loc[mask, "総合点"] >= threshold_high).astype(int)

        if y_sub.sum() < 5 or (len(y_sub) - y_sub.sum()) < 5:
            print("⚠️  特徴量重要度: ラベル偏りのためスキップ")
            return

        rf = RandomForestClassifier(n_estimators=300, random_state=42, n_jobs=-1)
        rf.fit(X_sub, y_sub)

        importances = pd.Series(rf.feature_importances_, index=use_cols).sort_values()

        # 日本語フォント設定（IPAGothicを優先）
        jp_fonts = [f.name for f in fm.fontManager.ttflist
                    if any(k in f.name for k in ["IPAGothic", "Noto", "Hiragino", "Yu Gothic"])]
        if jp_fonts:
            plt.rcParams["font.family"] = jp_fonts[0]

        fig, ax = plt.subplots(figsize=(8, 5))
        colors = ["#d9534f" if imp >= importances.quantile(0.67)
                  else "#5bc0de" if imp >= importances.quantile(0.33)
                  else "#aaaaaa" for imp in importances]
        importances.plot(kind="barh", ax=ax, color=colors)
        ax.set_title(f"特徴量重要度（{fiscal_year}年度）", fontsize=13)
        ax.set_xlabel("重要度")
        ax.axvline(importances.mean(), color="gray", linestyle="--", linewidth=0.8, label="平均")
        ax.legend(fontsize=9)
        plt.tight_layout()
        plt.savefig(FEATURE_IMPORTANCE_OUT, dpi=150)
        plt.close()
        print(f"✅  特徴量重要度PNG出力: {FEATURE_IMPORTANCE_OUT}")

    except Exception as e:
        print(f"⚠️  特徴量重要度出力エラー: {e}")


# ========= Final7選定 =========
def select_final7(df):
    """
    データ品質が「正常」の銘柄のみを対象に、
    総合点→安定性点→利回り点の順でソートし、
    同一業種最大2社の制限を適用して上位7社を返す。
    """
    df_valid  = df[df["データ品質"] == "正常"].copy()
    df_sorted = df_valid.sort_values(
        ["総合点", "安定性点", "利回り点"],
        ascending=[False, False, False]
    ).reset_index(drop=True)
 
    selected     = []
    sector_count = {}
 
    for _, row in df_sorted.iterrows():
        sector = row["業種"]
        if sector_count.get(sector, 0) < 2:
            selected.append(row)
            sector_count[sector] = sector_count.get(sector, 0) + 1
        if len(selected) == 7:
            break
 
    return pd.DataFrame(selected).reset_index(drop=True)
 
 
# ========= JPXデータ読み込み =========
print("JPXデータ読み込み中...")
jpx     = pd.read_excel(JPX_FILE)
prime   = jpx[jpx["市場・商品区分"].str.contains("プライム", na=False)]
tickers = (prime["コード"].astype(str) + ".T").tolist()
print(f"対象銘柄数: {len(tickers)}")
 
rows = []
 
# ==============================
# メイン処理（全銘柄スキャン）
# ==============================
for t in tqdm(tickers, desc="スキャン中"):
    try:
        stock = yf.Ticker(t)
        info  = stock.info
        div   = stock.dividends
 
        if div.empty:
            continue
 
        df_div     = div.to_frame(name="dividend")
        df_div["year"] = df_div.index.year
        yearly_div = df_div.groupby("year")["dividend"].sum().to_dict()
 
        # 15年連続配当チェック（会計年度ベース）
        if not all(yearly_div.get(y, 0) > 0 for y in range(start_year, fiscal_year)):
            continue
 
        # --- 【改良1】株価を履歴データから自前取得 ---
        current_price = get_current_price(stock)
        if current_price is None or current_price <= 0:
            for _key in ["currentPrice", "regularMarketPrice", "ask"]:
                _v = info.get(_key)
                if _v and float(_v) > 0:
                    current_price = float(_v)
                    break
            else:
                current_price = 0

        # --- 【改良2】配当利回りを直近3年平均配当で計算 ---
        avg_div_years  = [fiscal_year - 1 - i for i in range(AVG_YEARS)]
        avg_div_values = [yearly_div[y] for y in avg_div_years if yearly_div.get(y, 0) > 0]
        last_year_div  = yearly_div.get(fiscal_year - 1, 0)  # 前年確定値（表示用）

        # --- 【改良4】異常値フラグ（continueで除外しない） ---
        data_quality  = "正常"
        quality_notes = []

        if avg_div_values and current_price > 0:
            avg_div = sum(avg_div_values) / len(avg_div_values)
            if avg_div > current_price:
                # 配当 > 株価は物理的にあり得ない → dividendsデータ破損
                data_quality = "要確認"
                quality_notes.append(f"配当データ異常({avg_div:.0f}円)")
                dividend_yield = 0.0
            else:
                dividend_yield = (avg_div / current_price) * 100
        elif current_price > 0 and last_year_div > 0:
            if last_year_div > current_price:
                data_quality = "要確認"
                quality_notes.append(f"配当データ異常({last_year_div:.0f}円)")
                dividend_yield = 0.0
            else:
                dividend_yield = (last_year_div / current_price) * 100
        else:
            # info["dividendYield"]は信頼性が低いため使用しない
            dividend_yield = 0.0
 
        if dividend_yield <= 0 or dividend_yield > 15.0:
            data_quality = "要確認"
            quality_notes.append(f"利回り異常値({dividend_yield:.2f}%)")
            dividend_yield = 0.0
 
        if current_price <= 0:
            data_quality = "要確認"
            quality_notes.append("株価取得失敗")
 
        # --- 【改良5】株式分割チェック ---
        split_detected, split_note = check_recent_split(stock)
        if split_detected:
            data_quality = "要確認"
            quality_notes.append(f"株式分割あり({split_note})")
 
        # --- 【改良3】ROE・配当性向を複数年平均で計算 ---
        avg_roe, avg_payout = get_multi_year_financials(stock, AVG_YEARS)
 
        # NaN/None 両方をガード（NaN is not None → True のため個別チェックが必要）
        def _safe(val, fallback):
            if val is None:
                return fallback
            try:
                if math.isnan(float(val)):
                    return fallback
            except (TypeError, ValueError):
                pass
            return val
 
        roe    = _safe(avg_roe,    (info.get("returnOnEquity") or 0) * 100)
        payout = _safe(avg_payout, (info.get("payoutRatio")   or 0) * 100)
 
        if roe > 100 or roe < -100:
            quality_notes.append(f"ROE異常値({roe:.1f}%)")
            data_quality = "要確認"
            roe = 0.0
        if payout > 200 or payout < 0:
            quality_notes.append(f"配当性向異常値({payout:.1f}%)")
            data_quality = "要確認"
            payout = 0.0
        if roe == 0.0 and payout == 0.0:
            quality_notes.append("ROE取得失敗")
            data_quality = "要確認"
 
        growth = (info.get("revenueGrowth") or 0) * 100
        debt   =  info.get("debtToEquity")  or 0
 
        # --- 【改良8】DOE・理論利回り・実質PBR倍率の計算 ---
        # DOE(%) = ROE × 配当性向 / 100
        #   → 自己資本に対して何%配当しているかを示す。株価に依存しない指標。
        # 理論利回り(%) = DOE / PBR
        #   → 「あるべき利回り」の目安。現実利回りと比較することで
        #     株価が成長期待で先行していないかを判断できる。
        # 実質PBR倍率 = DOE / 現実利回り（≒実際のPBR）
        #   → 2倍超：成長期待で株価が先行、利回りは構造的に低くなりやすい
        #   → 1倍前後：株価が自己資本に対して適正水準
        pbr = info.get("priceToBook") or 0
        if roe > 0 and payout > 0:
            doe = round(roe * payout / 100, 2)
        else:
            doe = 0.0
        if pbr > 0 and doe > 0:
            theoretical_yield = round(doe / pbr, 2)
        else:
            theoretical_yield = 0.0
        if dividend_yield > 0 and doe > 0:
            pbr_ratio = round(doe / dividend_yield, 2)
        else:
            pbr_ratio = 0.0
 
        # スコアリング（各指標0〜10点）
        yield_score     = calc_score(dividend_yield, 1.0,   6.0)
        payout_score    = calc_score(payout,        20.0,  80.0, reverse=True)
        roe_score       = calc_score(roe,            3.0,  20.0)
        growth_score    = calc_score(growth,        -5.0,  15.0)
        debt_score      = calc_score(debt,           0.0, 300.0, reverse=True)
 
        # --- 【改良6】安定性スコアは分割逆換算済みdivで計算 ---
        yearly_div_normalized = normalize_div_for_stability(yearly_div, stock)
        stability_score = calc_dividend_stability(yearly_div_normalized, start_year, fiscal_year)
 
        # 総合点（最大100点）【C案】
        # 利回り点 ×2.5（25点）：利回り重視を強化
        # 安定性点 ×2.0（20点）：長期安定の根幹、据え置き
        # 配当性向点×2.0（20点）：配当持続性を強化（旧1.5→2.0）
        # ROE点    ×1.5（15点）：収益性、据え置き
        # 成長点   ×1.0（10点）：成長性、据え置き
        # 財務点   ×1.0（10点）：安全性、据え置き
        total_score = round(
            yield_score     * 2.5 +
            stability_score * 2.0 +
            payout_score    * 2.0 +
            roe_score       * 1.5 +
            growth_score    * 1.0 +
            debt_score      * 1.0,
            4
        )
 
        rows.append({
            "ティッカー":              t,
            "会社名":                  info.get("shortName", ""),
            "業種":                    info.get("sector", ""),
            "データ品質":              data_quality,
            "品質メモ":                " / ".join(quality_notes) if quality_notes else "",
            "株式分割(直近12M)":       split_note if split_detected else "",
            "株価":                    round(current_price, 0),
            "確定配当(前年)":          round(last_year_div, 2),
            "平均配当(3年)":           round(sum(avg_div_values)/len(avg_div_values), 2) if avg_div_values else 0,
            "利回り%(3年平均)":        round(dividend_yield, 2),
            "配当性向%(複数年平均)":   round(payout, 2),
            "ROE%(複数年平均)":        round(roe, 2),
            "DOE%":                    doe,
            "PBR":                     round(pbr, 2),
            "理論利回り%":             theoretical_yield,
            "実質PBR倍率":             pbr_ratio,
            "売上成長率%":             round(growth, 2),
            "負債比率":                round(debt, 2),
            "利回り点":                yield_score,
            "安定性点":                stability_score,
            "配当性向点":              payout_score,
            "ROE点":                   roe_score,
            "成長点":                  growth_score,
            "財務点":                  debt_score,
            "総合点":                  total_score,
            "15年連続配当":            "YES",
            "選定年":                  fiscal_year,
        })
 
    except Exception:
        pass
 
# ==============================
# DataFrame化・選定
# ==============================
df_all = pd.DataFrame(rows)
 
if df_all.empty:
    print("⚠️  候補銘柄が0件です。JPXファイルやネット接続を確認してください。")
    exit()

# ==============================
# 【追加】異常検知 & 特徴量重要度
# ==============================
print("\n異常検知を実行中...")
df_all = detect_anomalies(df_all)

if EXPORT_FEATURE_IMPORTANCE:
    print("特徴量重要度を分析中...")
    export_feature_importance(df_all)
 
df_final = select_final7(df_all)
df_check = df_all[df_all["データ品質"] == "要確認"].sort_values("総合点", ascending=False)

# 異常検知結果（正常銘柄の中でフラグ=-1のもの）
df_anomaly = df_all[
    (df_all["データ品質"] == "正常") & (df_all["anomaly_flag"] == -1)
].sort_values("anomaly_score", ascending=True)

# ==============================
# Excel出力（4シート構成）
# ==============================
with pd.ExcelWriter(ANNUAL_RESULT_FILE, engine="openpyxl") as writer:
    df_final.to_excel(
        writer, index=False, sheet_name="Final7"
    )
    df_all[df_all["データ品質"] == "正常"].sort_values(
        "総合点", ascending=False
    ).to_excel(writer, index=False, sheet_name="AllCandidates")
    if not df_check.empty:
        df_check.to_excel(writer, index=False, sheet_name="DataCheck_要確認")
    if not df_anomaly.empty:
        anomaly_cols = [
            "ティッカー", "会社名", "業種",
            "利回り%(3年平均)", "配当性向%(複数年平均)", "ROE%(複数年平均)",
            "DOE%", "PBR", "理論利回り%", "実質PBR倍率",
            "売上成長率%", "負債比率", "総合点",
            "anomaly_score", "anomaly_reason",
        ]
        out_cols = [c for c in anomaly_cols if c in df_anomaly.columns]
        df_anomaly[out_cols].to_excel(writer, index=False, sheet_name="AnomalyReport")

# ==============================
# 【改良9】Final7シート：異常銘柄を黄色背景・赤太字でハイライト
# ==============================
if "anomaly_flag" in df_final.columns and df_final["anomaly_flag"].eq(-1).any():
    from openpyxl import load_workbook
    wb = load_workbook(ANNUAL_RESULT_FILE)
    ws = wb["Final7"]
    WARN_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
    headers = {cell.value: cell.column for cell in ws[1]}
    flag_col = headers.get("anomaly_flag")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if flag_col and row[flag_col - 1].value == -1:
            for cell in row:
                cell.fill = WARN_FILL
                cell.font = Font(bold=True, color="CC0000")
    wb.save(ANNUAL_RESULT_FILE)
    print("⚠️  Final7シート：異常銘柄を黄色警告表示しました")

# ==============================
# Final7をJSONに保存
# ==============================
final7_data = {
    "selected_date": today.strftime("%Y-%m-%d"),
    "tickers": df_final["ティッカー"].tolist(),
    "details": df_final.to_dict(orient="records")
}
with open(FINAL7_JSON, "w", encoding="utf-8") as f:
    json.dump(final7_data, f, ensure_ascii=False, indent=2)
 
# ==============================
# 結果表示・メール通知
# ==============================
print("\n" + "=" * 60)
print(f"【年次選定完了】{today.strftime('%Y年%m月%d日')}")
print(f"全候補: {len(df_all)} 社  /  Final7: {len(df_final)} 社  /  要確認: {len(df_check)} 社  /  異常検知: {len(df_anomaly)} 社")
print("=" * 60)
print(df_final[["ティッカー", "会社名", "利回り%(3年平均)", "DOE%", "理論利回り%", "実質PBR倍率", "安定性点", "総合点"]].to_string(index=False))

if not df_check.empty:
    print(f"\n⚠️  データ要確認銘柄: {len(df_check)} 社")
    print("→ annual_result.xlsx の「DataCheck_要確認」シートを確認してください")

if not df_anomaly.empty:
    print(f"\n🔍  異常検知銘柄: {len(df_anomaly)} 社（指標に統計的な乖離あり）")
    print("→ annual_result.xlsx の「AnomalyReport」シートを確認してください")
    disp_cols = ["ティッカー", "会社名", "anomaly_score", "anomaly_reason"]
    disp_cols = [c for c in disp_cols if c in df_anomaly.columns]
    print(df_anomaly[disp_cols].head(10).to_string(index=False))

print(f"\n結果保存: {ANNUAL_RESULT_FILE}")
print(f"監視用JSON: {FINAL7_JSON}")
if EXPORT_FEATURE_IMPORTANCE and os.path.exists(FEATURE_IMPORTANCE_OUT):
    print(f"特徴量重要度PNG: {FEATURE_IMPORTANCE_OUT}")

anomaly_section = ""
if not df_anomaly.empty:
    disp_cols = ["ティッカー", "会社名", "anomaly_score", "anomaly_reason"]
    disp_cols = [c for c in disp_cols if c in df_anomaly.columns]
    anomaly_section = (
        f"\n\n🔍 異常検知銘柄: {len(df_anomaly)} 社\n"
        f"{df_anomaly[disp_cols].head(10).to_string(index=False)}"
    )

body = (
    f"【年次選定完了】{today.strftime('%Y年%m月%d日')}\n\n"
    f"全候補: {len(df_all)} 社 / Final7: {len(df_final)} 社 / 要確認: {len(df_check)} 社 / 異常検知: {len(df_anomaly)} 社\n\n"
    f"{df_final[['ティッカー','会社名','利回り%(3年平均)','DOE%','理論利回り%','実質PBR倍率','総合点']].to_string(index=False)}\n\n"
    + (f"⚠️ データ要確認銘柄あり: {len(df_check)} 社\n"
       f"{df_check[['ティッカー','会社名','品質メモ']].to_string(index=False)}"
       if not df_check.empty else "")
    + anomaly_section
)
send_alert(f"【配当システム】{fiscal_year}年度 年次選定結果", body)
print("\n完了。")
 

